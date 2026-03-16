# eproc.py
# Funções específicas para automação do sistema e-Proc

from .navegador import obter_driver, minimizar_navegador
from .logger import logger
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.support.ui import Select
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.common.action_chains import ActionChains
from selenium.common.exceptions import TimeoutException, NoSuchElementException, NoAlertPresentException
from tkinter import filedialog
from openpyxl import load_workbook
import time
import re
import threading

# Dicionário de códigos de erro conhecidos do EPROC
CODIGOS_ERRO_EPROC = {
    0: "Processo já Migrado",
    1: "Erro INESPERADO NA MIGRAÇÃO",
    5: "Processo NÃO CONSTA na base do SAJ",
    6: "Migração NÃO HABILITADA para este processo",
    13: "JUÍZO NÃO HABILITADO PARA MIGRAÇÃO",
    14: "COMPETÊNCIA não cadastrada no EPROC",
    20: "Existem MANDADOS PENDENTES",
    23: "Existem ARs sem Devolução",
    25: "Erros nas peças processuais",
    34: "MIGRAÇÃO NÃO PERMITIDA",
    35: "Erro no Cadastro de Parte",
    36: "Inconsistência no Cadastro de Parte",
    47: "Advogado ou Escritório não cadastrado no EPROC",
    51: "Processo Será Migrado Após Solução do Incidente/Apenso",
    60: "Migração NÃO HABILITADA, Processo Entranhado",
    61: "Processo em Segundo Grau - Agravo/Recurso",
    66: "Existem ARs Ag. Envio aos Correios",
    89: "Existem mandados pendentes que ainda não constam na fila e situação: Ag. Cumprimento pelo Oficial",
    95: "PROCESSO BAIXADO NÃO SERÁ MIGRADO",
    96: "CNPJ NÃO INFORMADO",    
        
}

# Timeout (em segundos) para aguardar o resultado da migração após clicar em Migrar.
# Opções: 180 (3 min), 300 (5 min), 540 (9 min).
# Menor = mais rápido no início (processa os simples primeiro; os lentos ficam para retry).
# Maior = espera mais tempo por processos complexos antes de dar timeout. Processos com mais páginas demoram mais.
TIMEOUT_MIGRACAO = 300  # 5 minutos (opções: 180, 300, 540)

# Controle de cancelamento da execução (registrado pela UI)
_cancel_event: threading.Event | None = None


def registrar_cancelamento(evento_cancelamento: threading.Event):
    global _cancel_event
    _cancel_event = evento_cancelamento


def _cancelado() -> bool:
    return _cancel_event is not None and _cancel_event.is_set()


def _tentar_recuperar_navegador(driver) -> bool:
    """
    Tenta recuperar o navegador após erro/timeout recarregando a página.
    Retorna True se conseguiu recuperar, False caso contrário.
    """
    try:
        logger.info("Recarregando a página para recuperar...")
        driver.refresh()
        time.sleep(3)
        # Clicar no atalho após recarregar
        elemento_atalho = WebDriverWait(driver, 10).until(
            EC.element_to_be_clickable((By.CSS_SELECTOR, '#atalhoRapido_2'))
        )
        elemento_atalho.click()
        logger.info("✅ Página recarregada! Continuando com o próximo processo...")
        return True
    except Exception:
        return False


def extrair_codigo_erro(driver):
    """
    Extrai todos os códigos de erro únicos exibidos na seção "Informações da Migração para o EPROC".
    - 1º tenta ler diretamente os <b> de cada <li> via XPath absoluto mostrado pelo usuário
    - 2º faz fallback lendo a célula (td) após o cabeçalho "Erros:"
    Retorna lista [(codigo:int, mensagem:str), ...] sem repetição
    """
    erros_unicos = {}

    def registrar_codigo_e_mensagem(texto: str) -> None:
        # Normalizar aspas e espaços; pegar apenas a 1ª linha, que contém o código
        if not texto:
            return
        texto_normalizado = texto.replace("“", '"').replace("”", '"').strip()
        primeira_linha = texto_normalizado.splitlines()[0].strip()
        # Aceita aspas e marcador no início
        match = re.search(r'^[\s\u2022\"\']*(\d+)\s*-\s*(.+)$', primeira_linha)
        if match:
            codigo = int(match.group(1))
            mensagem = match.group(2).strip()
            if codigo not in erros_unicos:
                erros_unicos[codigo] = mensagem

    try:
        # 1) XPath direto para os <b> dentro da lista de erros (fieldset[3] - pós-migração)
        xpath_b = "//*[@id='frmProcessoCadastro']/fieldset[3]//ul/li/b"
        elementos_b = driver.find_elements(By.XPATH, xpath_b)
        for el in elementos_b:
            registrar_codigo_e_mensagem(el.text)
    except Exception:
        pass

    if not erros_unicos:
        try:
            # 2) Fallback: ler a célula que segue o cabeçalho "Erros:" e processar cada linha
            xpath_container = "//th[contains(text(), 'Erros:')]/following-sibling::td"
            container_erro = driver.find_element(By.XPATH, xpath_container)
            texto_completo = container_erro.text
            for linha in texto_completo.split('\n'):
                registrar_codigo_e_mensagem(linha)
        except Exception:
            pass

    if not erros_unicos:
        try:
            # 3) Fieldsets cujo legend contenha "Erro" (ex: "Erros de CNPJ")
            #    Captura o texto completo de cada <li> para incluir texto em tags aninhadas
            xpath_li = "//fieldset[legend[contains(text(), 'Erro')]]//li"
            elementos_li = driver.find_elements(By.XPATH, xpath_li)
            for el in elementos_li:
                registrar_codigo_e_mensagem(el.text)
        except Exception:
            pass

    return list(erros_unicos.items())


def obter_descricao_erro(codigo):
    """
    Retorna a descrição amigável do código de erro, se conhecida.
    """
    if codigo in CODIGOS_ERRO_EPROC:
        return CODIGOS_ERRO_EPROC[codigo]
    else:
        return f"Erro código {codigo}"


def aguardar_sucesso_ou_erros(driver, timeout: int = 180):
    """
    Após acionar a migração, aguarda até que apareça:
    - uma indicação de sucesso (texto contendo "Sucesso"), ou
    - erros listados na seção de informações.

    Retorna uma tupla (status, detalhes):
      status ∈ {"sucesso", "erro", "timeout", "cancelado"}
      detalhes: lista de (codigo:int, mensagem:str) quando status=="erro"
    """
    momento_limite = time.time() + timeout
    while time.time() < momento_limite:
        if _cancelado():
            return ("cancelado", [])
        # Verifica sucesso
        try:
            driver.find_element(By.XPATH, '//*[contains(text(), "Sucesso")]')
            return ("sucesso", [])
        except Exception:
            pass
        # Verifica erros
        erros_detectados = extrair_codigo_erro(driver)
        if erros_detectados:
            return ("erro", erros_detectados)
        # Verifica exceção de validação do sistema (ex: "processo atual não coincide com o último extraído")
        try:
            div_excecao = driver.find_element(By.CSS_SELECTOR, '#divInfraExcecao')
            if div_excecao.is_displayed():
                try:
                    span = div_excecao.find_element(By.CSS_SELECTOR, '.infraExcecao')
                    msg = span.text.strip()
                except Exception:
                    msg = div_excecao.text.strip()
                if msg:
                    return ("excecao", msg)
        except Exception:
            pass
        time.sleep(0.5)
    return ("timeout", [])


def tratar_pessoas_sem_cpf(driver):
    """
    Identifica selects na tabela de pessoas sem CPF e seleciona a opção 'Sem CPF' (value='3').
    Após selecionar, marca o checkbox de declaração necessário para habilitar a migração.
    Retorna True se realizou as seleções, False caso contrário.
    """
    try:
        # Verificar se a seção existe
        xpath_secao = '//*[contains(text(), "Pessoas sem CPF/CNPJ")]'
        if not driver.find_elements(By.XPATH, xpath_secao):
            return False

        logger.warning("⚠️ Detectado 'Pessoas sem CPF/CNPJ'. Tentando selecionar opção 'Sem CPF'...")
        
        # Encontrar todos os selects dentro da tabela de pessoas sem CPF
        # Localiza o fieldset ou container
        try:
            # Tenta encontrar o fieldset específico
            container = driver.find_element(By.XPATH, "//fieldset[legend[contains(text(), 'Pessoas sem CPF/CNPJ')]]")
        except:
            # Se não encontrar fieldset, tenta encontrar a tabela diretamente
            try:
                container = driver.find_element(By.XPATH, "//legend[contains(text(), 'Pessoas sem CPF/CNPJ')]/following-sibling::table")
            except:
                # Última tentativa: procura qualquer elemento que contenha o texto e pega o parent
                elemento_secao = driver.find_element(By.XPATH, xpath_secao)
                container = elemento_secao.find_element(By.XPATH, "./ancestor::fieldset | ./following-sibling::table")
        
        selects = container.find_elements(By.TAG_NAME, "select")
        
        if not selects:
            logger.warning("⚠️ Nenhum dropdown encontrado na tabela.")
            return False
            
        count = 0
        for select_element in selects:
            try:
                # Verificar se o select está visível e habilitado
                if not select_element.is_displayed() or not select_element.is_enabled():
                    continue
                    
                select_obj = Select(select_element)
                # Selecionar a opção "Parte SEM CPF" pelo texto visível
                opcao_sem_cpf = next(
                    (opt for opt in select_obj.options if "SEM CPF" in opt.text.upper()),
                    None
                )
                if opcao_sem_cpf:
                    select_obj.select_by_visible_text(opcao_sem_cpf.text)
                    logger.info(f"✅ Selecionado '{opcao_sem_cpf.text}' para o dropdown {count + 1}")
                else:
                    # Fallback: seleciona pelo value "3" caso o texto mude
                    select_obj.select_by_value("3")
                    logger.info(f"✅ Selecionado 'Parte SEM CPF' (value=3) para o dropdown {count + 1}")
                count += 1
                
                # Pequena pausa para o evento onchange processar
                time.sleep(0.5) 
            except Exception as e:
                logger.warning(f"⚠️ Erro ao selecionar opção no dropdown {count + 1}: {e}")
        
        if count > 0:
            logger.info(f"✅ Total de {count} dropdown(s) configurado(s) com 'Sem CPF'")
            
            # Logo após selecionar, o checkbox de declaração aparece
            # Aguardar um momento para o checkbox aparecer
            time.sleep(1)
            
            # Procurar e marcar o checkbox de declaração
            logger.info("🔍 Procurando checkbox de declaração...")
            try:
                checkbox = None
                
                # Tenta diferentes formas de localizar o checkbox (do mais específico ao mais genérico)
                try:
                    # Método 1: ID específico do checkbox (mais confiável)
                    checkbox = driver.find_element(By.ID, "chk_ciencia_pessoa_sem_cpf")
                    logger.info("✅ Checkbox encontrado pelo ID 'chk_ciencia_pessoa_sem_cpf'")
                except:
                    try:
                        # Método 2: Por atributos do input
                        checkbox = driver.find_element(By.XPATH, 
                            "//input[@type='checkbox' and (contains(@id, 'ciencia') or contains(@id, 'cpf') or contains(@id, 'pessoa'))]")
                        logger.info("✅ Checkbox encontrado por atributos")
                    except:
                        try:
                            # Método 3: Checkbox próximo ao texto da declaração
                            checkbox = driver.find_element(By.XPATH, 
                                "//input[@type='checkbox' and ancestor::*[contains(text(), 'Para as pessoas físicas Sem o CPF')]]")
                            logger.info("✅ Checkbox encontrado próximo ao texto de declaração")
                        except:
                            # Método 4: Último recurso - qualquer checkbox no container
                            checkbox = driver.find_element(By.XPATH, 
                                "//*[contains(text(), 'Pessoas sem CPF/CNPJ')]/ancestor::fieldset//input[@type='checkbox']")
                            logger.info("✅ Checkbox encontrado no container")
                
                if checkbox:
                    # Verificar se está visível e habilitado
                    if checkbox.is_displayed() and checkbox.is_enabled():
                        # Verificar se já está marcado
                        if not checkbox.is_selected():
                            # Clicar no checkbox
                            checkbox.click()
                            logger.info("✅ Checkbox de declaração marcado com sucesso!")
                            time.sleep(1)
                        else:
                            logger.info("✅ Checkbox de declaração já estava marcado.")
                        return True
                    else:
                        logger.warning("⚠️ Checkbox encontrado mas não está visível ou habilitado.")
                        return True
                else:
                    logger.warning("⚠️ Checkbox de declaração não encontrado.")
                    return True
                    
            except Exception as e:
                logger.warning(f"⚠️ Erro ao tentar marcar checkbox de declaração: {e}")
                logger.warning("⚠️ Tentando clicar via JavaScript como alternativa...")
                try:
                    # Tentativa via JavaScript
                    driver.execute_script("document.getElementById('chk_ciencia_pessoa_sem_cpf').checked = true;")
                    logger.info("✅ Checkbox marcado via JavaScript!")
                    time.sleep(1)
                    return True
                except Exception as e2:
                    logger.error(f"❌ Falha ao marcar via JavaScript: {e2}")
                    return True  # Retorna True mesmo assim
        else:
            return False

    except Exception as e:
        logger.error(f"❌ Erro ao tratar pessoas sem CPF: {e}")
        return False


def acessar_localizadores():
    """Acessa a seção 'Meus Localizadores' no e-Proc"""
    driver = obter_driver()
    if driver is None:
        logger.warning("Navegador não está pronto. Acesse o e-Proc primeiro.")
        return

    try:
        selector = 'i[title="Meus Localizadores"]'
        logger.info("Acessando 'Meus Localizadores'...")
        elemento = WebDriverWait(driver, 5).until(
            EC.element_to_be_clickable((By.CSS_SELECTOR, selector))
        )
        elemento.click()
        logger.info("'Meus Localizadores' acessado com sucesso.")
    except Exception as e:
        logger.error(f"Erro ao tentar acessar 'Meus Localizadores': {e}")


def migrador():
    """Abre diálogo para selecionar arquivo Excel e processa os processos"""
    
    # Abrir diálogo para selecionar arquivo Excel ANTES de acessar o navegador
    logger.info("Selecione o arquivo Excel com a lista de processos...")
    arquivo_excel = filedialog.askopenfilename(
        title="Selecione o arquivo Excel",
        filetypes=[("Arquivos Excel", "*.xlsx *.xls"), ("Todos os arquivos", "*.*")]
    )
    
    if not arquivo_excel:
        logger.info("Nenhum arquivo selecionado. Operação cancelada.")
        return
    
    try:
        # Ler o arquivo Excel com openpyxl
        logger.info(f"Lendo arquivo: {arquivo_excel}")
        wb = load_workbook(arquivo_excel)
        ws = wb['Planilha1']
        
        # Encontrar as colunas 'Processo' e 'Status' na primeira linha
        coluna_processo = None
        coluna_status = None

        for col_idx, cell in enumerate(ws[1], 1):
            if cell.value:
                valor_celula = str(cell.value).strip()
                if valor_celula == 'Processo':
                    coluna_processo = col_idx
                elif valor_celula == 'Status':
                    coluna_status = col_idx

        if coluna_processo is None:
            logger.info("Erro: Coluna 'Processo' não encontrada no arquivo Excel.")
            wb.close()
            return

        # Se a coluna Status não existir, criar na 2ª coluna
        if coluna_status is None:
            coluna_status = 2
            ws.cell(row=1, column=coluna_status, value="Status")
            wb.save(arquivo_excel)
            logger.info(f"✅ Coluna 'Status' criada na coluna {coluna_status}.")

        # Obter lista de processos (ignorando a primeira linha - cabeçalho)
        dados_processos = []
        for row_idx, row in enumerate(ws.iter_rows(min_row=2), start=2):
            valor_processo = row[coluna_processo - 1].value
            if valor_processo:
                proc_str = str(valor_processo).strip()
                if '/' in proc_str:
                    proc_str, incidente_str = proc_str.rsplit('/', 1)
                    incidente_str = incidente_str.strip() or '0'
                else:
                    incidente_str = '0'
                dados_processos.append({
                    'processo': proc_str,
                    'incidente': incidente_str,
                    'linha': row_idx  # Guardar o número da linha para atualizar depois
                })
        
        total_processos = len(dados_processos)
        logger.info(f"Total de processos encontrados: {total_processos}")
        
        # AGORA verifica se o navegador está pronto
        driver = obter_driver()
        if driver is None:
            logger.info("Navegador não está pronto. Acesse o e-Proc primeiro.")
            wb.close()
            return
        
        # Minimizar a janela do navegador após início da execução
        minimizar_navegador()
        
        # Clicar no atalhoRapido_2
        try:
            selector = '#atalhoRapido_2'
            logger.info("Clicando em 'atalhoRapido_2'...")
            elemento = WebDriverWait(driver, 5).until(
                EC.element_to_be_clickable((By.CSS_SELECTOR, selector))
            )
            elemento.click()
            logger.info("Elemento 'atalhoRapido_2' clicado com sucesso.")
        except Exception as e:
            logger.error(f"❌ ERRO: Atalho de migração não encontrado no e-Proc. Verifique se está na página correta. ({e})")
            wb.close()
            return
        
        # Iterar pelos processos
        for i, dados in enumerate(dados_processos, 1):
            if _cancelado():
                logger.info("⚠️ Execução cancelada pelo usuário.")
                break
            processo = dados['processo']
            incidente = dados['incidente']
            linha = dados['linha']
            logger.info(f"\n--- [{i}/{total_processos}] Processando: {processo} (Incidente: {incidente}) ---")
            
            try:
                # Preencher campo do número do processo
                campo_processo = '#txtNumProcesso'
                elemento = WebDriverWait(driver, 5).until(
                    EC.presence_of_element_located((By.CSS_SELECTOR, campo_processo))
                )
                elemento.clear()
                elemento.send_keys(processo)
                logger.info(f"Número do processo '{processo}' inserido.")
                
                # Preencher campo do número do incidente
                campo_incidente = '#txtNumSeqProcesso'
                elemento = driver.find_element(By.CSS_SELECTOR, campo_incidente)
                elemento.clear()
                elemento.send_keys(incidente)
                logger.info(f"Número do incidente '{incidente}' inserido.")
                
                # Clicar no botão "Buscar Processo para Migração"
                botao_buscar = '#btnBuscar'
                elemento = driver.find_element(By.CSS_SELECTOR, botao_buscar)
                elemento.click()
                logger.info(f"Botão 'Buscar Processo para Migração' clicado.")
                for _ in range(10):
                    if _cancelado():
                        break
                    time.sleep(0.1)
                if _cancelado():
                    logger.info("Cancelado após clique em Buscar.")
                    break
                
                # Aguardar a seção "Dados do Processo" carregar
                logger.info("Aguardando carregamento dos dados do processo (até 5 minutos)...")
                try:
                    # Aguardar o elemento "Dados do Processo" aparecer
                    WebDriverWait(driver, 300).until(
                        EC.visibility_of_element_located((By.XPATH, '//legend[contains(text(), "Dados do Processo")]'))
                    )
                    logger.info("✅ Dados do processo carregados.")
                    for _ in range(10):
                        if _cancelado():
                            break
                        time.sleep(0.1)
                    if _cancelado():
                        logger.info("Cancelado após carregar dados do processo.")
                        break
                    
                    # Agora verificar se tem botão Migrar ou mensagem de erro
                    tem_botao_migrar = False
                    tem_erro = False
                    tem_pessoas_sem_cpf = False
                    
                    try:
                        elemento = driver.find_element(By.CSS_SELECTOR, '#btnMigrar')
                        tem_botao_migrar = elemento.is_displayed()
                    except:
                        pass
                    
                    # Detectar erros pela extração direta dos códigos na seção de erros
                    erros_lista = extrair_codigo_erro(driver)
                    tem_erro = len(erros_lista) > 0
                    
                    try:
                        elemento = driver.find_element(By.XPATH, '//*[contains(text(), "Pessoas sem CPF/CNPJ")]')
                        tem_pessoas_sem_cpf = elemento.is_displayed()
                    except:
                        pass
                    
                    # Decidir o que fazer baseado no que encontrou
                    if tem_pessoas_sem_cpf:
                        # Encontrado "Pessoas sem CPF/CNPJ" - não pode migrar
                        logger.info("❌ Erro encontrado: 'Pessoas sem CPF/CNPJ'")
                        
                        # Atualizar status na planilha
                        ws.cell(row=linha, column=coluna_status, value="Não Migrado SEM CPF-CNPJ")
                        wb.save(arquivo_excel)
                        logger.info(f"Status atualizado na planilha: Não Migrado SEM CPF-CNPJ")
                        
                        # Clicar novamente no link para o próximo processo
                        logger.info("Clicando novamente em 'atalhoRapido_2' para o próximo processo...")
                        elemento = WebDriverWait(driver, 5).until(
                            EC.element_to_be_clickable((By.CSS_SELECTOR, '#atalhoRapido_2'))
                        )
                        elemento.click()
                        logger.info("Pulando para o próximo processo...")
                        continue
                        
                    elif tem_botao_migrar:
                        logger.info("✅ Botão 'Migrar Processo' encontrado.")
                        
                        # Executar a função JavaScript diretamente
                        logger.info("Executando função 'importarNovo()'...")
                        driver.execute_script("importarNovo()")
                        
                        # Aguardar 2 segundos antes de aceitar o alert
                        logger.info("Aguardando caixa de diálogo de confirmação...")
                        for _ in range(20):
                            if _cancelado():
                                break
                            time.sleep(0.1)
                        if _cancelado():
                            logger.info("Cancelado antes de aceitar o alerta.")
                            break
                        
                        # Aceitar o alert JavaScript
                        alert = driver.switch_to.alert
                        alert.accept()
                        logger.info("✅ Alert JavaScript aceito.")
                        for _ in range(10):
                            if _cancelado():
                                break
                            time.sleep(0.1)
                        if _cancelado():
                            logger.info("Cancelado após aceitar o alerta.")
                            break
                        
                        # Aguardar resultado (sucesso ou erros)
                        logger.info("Aguardando resultado da migração (sucesso ou erros)...")
                        # timeout: 180 (3 min), 300 (5 min), 540 (9 min) - configurável em TIMEOUT_MIGRACAO
                        status_resultado, erros_pos = aguardar_sucesso_ou_erros(driver, timeout=TIMEOUT_MIGRACAO)

                        if status_resultado == "sucesso":
                            logger.info("✅ Processo migrado com sucesso!")
                            # Atualizar status na planilha
                            ws.cell(row=linha, column=coluna_status, value="Migrado com Sucesso")
                            wb.save(arquivo_excel)
                            logger.info(f"Status atualizado na planilha: Migrado com Sucesso")
                            # Clicar novamente no link para o próximo processo
                            logger.info("Clicando novamente em 'atalhoRapido_2' para o próximo processo...")
                            elemento = WebDriverWait(driver, 5).until(
                                EC.element_to_be_clickable((By.CSS_SELECTOR, '#atalhoRapido_2'))
                            )
                            elemento.click()
                        elif status_resultado == "erro":
                            logger.info("❌ Erros pós-migração detectados.")
                            descricoes = []
                            for codigo, mensagem in erros_pos:
                                descricao = mensagem if mensagem else obter_descricao_erro(codigo)
                                descricoes.append(f"Erro {codigo}: {descricao}")
                                logger.info(f"   Código identificado: {codigo} - {descricao}")
                            status_texto = "; ".join(descricoes) if descricoes else "Não Migrado - Com Erros"
                            # Atualizar status na planilha
                            ws.cell(row=linha, column=coluna_status, value=status_texto)
                            wb.save(arquivo_excel)
                            logger.info(f"Status atualizado na planilha: {status_texto}")
                            # Tentar re-clicar no botão Migrar para forçar refresh
                            try:
                                logger.info("Reclicando no botão 'Migrar Processo' para forçar refresh...")
                                botao_migrar_refresh = WebDriverWait(driver, 5).until(
                                    EC.element_to_be_clickable((By.CSS_SELECTOR, '#btnMigrar'))
                                )
                                botao_migrar_refresh.click()
                                # Se abrir o alerta de confirmação novamente, aceitar
                                try:
                                    WebDriverWait(driver, 5).until(EC.alert_is_present())
                                    driver.switch_to.alert.accept()
                                    logger.info("Alerta de confirmação pós-erro aceito.")
                                except TimeoutException:
                                    pass
                                # Pequena espera e tentativa de voltar pelo atalho para garantir refresh
                                for _ in range(10):
                                    if _cancelado():
                                        break
                                    time.sleep(0.1)
                                try:
                                    elemento = WebDriverWait(driver, 5).until(
                                        EC.element_to_be_clickable((By.CSS_SELECTOR, '#atalhoRapido_2'))
                                    )
                                    elemento.click()
                                except Exception:
                                    pass
                            except Exception:
                                logger.info("Não foi possível re-clicar em 'Migrar Processo'. Voltando pelo atalho...")
                                try:
                                    elemento = WebDriverWait(driver, 5).until(
                                        EC.element_to_be_clickable((By.CSS_SELECTOR, '#atalhoRapido_2'))
                                    )
                                    elemento.click()
                                except Exception:
                                    pass
                        elif status_resultado == "excecao":
                            logger.info(f"❌ Exceção do sistema detectada: {erros_pos}")
                            ws.cell(row=linha, column=coluna_status, value=f"ERRO SISTEMA: {erros_pos}")
                            wb.save(arquivo_excel)
                            logger.info(f"Status atualizado na planilha: ERRO SISTEMA: {erros_pos}")
                            try:
                                elemento = WebDriverWait(driver, 5).until(
                                    EC.element_to_be_clickable((By.CSS_SELECTOR, '#atalhoRapido_2'))
                                )
                                elemento.click()
                            except Exception:
                                pass
                            continue
                        elif status_resultado == "timeout":
                            logger.info("❌ TIMEOUT: Sistema não respondeu em 5 minutos após migração.")
                            logger.info("⚠️ Tentando recuperar e continuar com o próximo processo...")
                            ws.cell(row=linha, column=coluna_status, value="ERRO SISTEMA: Timeout após migração")
                            wb.save(arquivo_excel)
                            logger.info("⚠️ Possível problema de conexão, sistema lento ou travado.")

                            # Tentar recuperar o navegador
                            if _tentar_recuperar_navegador(driver):
                                continue
                            else:
                                logger.info("❌ Falha na recuperação.")
                                logger.info("⚠️ EXECUÇÃO INTERROMPIDA - Não foi possível recuperar o navegador.")
                                break
                        else:  # cancelado
                            logger.info("Cancelado durante a espera pelo resultado da migração.")
                            break

                    elif tem_erro:
                        # Erros de validação encontrados via extração direta
                        logger.info("❌ Erros de validação encontrados.")
                        if erros_lista:
                            descricoes = []
                            for codigo, mensagem in erros_lista:
                                descricao = mensagem if mensagem else obter_descricao_erro(codigo)
                                descricoes.append(f"Erro {codigo}: {descricao}")
                                logger.info(f"   Código identificado: {codigo} - {descricao}")
                            status_texto = "; ".join(descricoes)
                        else:
                            status_texto = "Não Migrado - Com Erros"

                        # Atualizar status na planilha
                        ws.cell(row=linha, column=coluna_status, value=status_texto)
                        wb.save(arquivo_excel)
                        logger.info(f"Status atualizado na planilha: {status_texto}")

                        # Clicar novamente no link para o próximo processo
                        logger.info("Clicando novamente em 'atalhoRapido_2' para o próximo processo...")
                        elemento = WebDriverWait(driver, 5).until(
                            EC.element_to_be_clickable((By.CSS_SELECTOR, '#atalhoRapido_2'))
                        )
                        elemento.click()
                        logger.info("Pulando para o próximo processo...")
                        continue

                    else:
                        # Nem botão nem erro encontrado
                        logger.info("⚠️ Nem botão Migrar nem mensagem de erro foram encontrados.")

                        # Atualizar status na planilha
                        ws.cell(row=linha, column=coluna_status, value="Não Migrado - Com Erros")
                        wb.save(arquivo_excel)
                        logger.info(f"Status atualizado na planilha: Não Migrado - Com Erros")

                        # Clicar novamente no link para o próximo processo
                        logger.info("Clicando novamente em 'atalhoRapido_2' para o próximo processo...")
                        try:
                            elemento = WebDriverWait(driver, 5).until(
                                EC.element_to_be_clickable((By.CSS_SELECTOR, '#atalhoRapido_2'))
                            )
                            elemento.click()
                        except:
                            pass
                        logger.info("Pulando para o próximo processo...")
                        continue

                except TimeoutException as e:
                    # Timeout ao carregar dados do processo - tentar recuperar
                    logger.info(f"❌ TIMEOUT: Dados do processo não carregaram em 5 minutos.")
                    logger.info(f"⚠️ Tentando recuperar e continuar com o próximo processo...")
                    logger.info(f"⚠️ Detalhes: {e}")
                    
                    # Atualizar status na planilha
                    ws.cell(row=linha, column=coluna_status, value="ERRO SISTEMA: Timeout ao carregar dados")
                    wb.save(arquivo_excel)
                    
                    # Tentar recuperar o navegador
                    if _tentar_recuperar_navegador(driver):
                        continue
                    else:
                        logger.info("❌ Falha na recuperação.")
                        logger.info("⚠️ EXECUÇÃO INTERROMPIDA - Não foi possível recuperar o navegador.")
                        break
                    
                except Exception as e:
                    # Outro erro inesperado - tentar recuperar
                    logger.info(f"❌ ERRO ao processar dados do processo: {e}")
                    logger.info(f"⚠️ Tentando recuperar e continuar com o próximo processo...")
                    
                    # Atualizar status na planilha
                    ws.cell(row=linha, column=coluna_status, value="ERRO SISTEMA: Erro ao processar dados")
                    wb.save(arquivo_excel)
                    
                    # Tentar recuperar o navegador
                    if _tentar_recuperar_navegador(driver):
                        continue
                    else:
                        logger.info("❌ Falha na recuperação.")
                        logger.info("⚠️ EXECUÇÃO INTERROMPIDA - Não foi possível recuperar o navegador.")
                        break
                
            except Exception as e:
                # Erro ao preencher campos, clicar em botões - tentar recuperar
                logger.info(f"❌ ERRO ao processar o processo {processo}: {e}")
                logger.info(f"⚠️ Tentando recuperar e continuar com o próximo processo...")
                logger.info(f"⚠️ Possíveis causas: elementos não encontrados, conexão temporária perdida.")
                
                # Atualizar status na planilha
                ws.cell(row=linha, column=coluna_status, value="ERRO SISTEMA: Falha ao processar")
                wb.save(arquivo_excel)
                
                # Tentar recuperar o navegador
                if _tentar_recuperar_navegador(driver):
                    continue
                else:
                    logger.info("❌ Falha na recuperação.")
                    logger.info("⚠️ EXECUÇÃO INTERROMPIDA - Não foi possível recuperar o navegador.")
                    break
        
        # Fechar a planilha no final
        wb.close()
        logger.info(f"\n✅ Processamento concluído! Planilha atualizada: {arquivo_excel}")
            
    except Exception as e:
        logger.info(f"Erro ao processar arquivo Excel: {e}")


def migrador_sem_cpf():
    """
    Abre diálogo para selecionar arquivo Excel e processa os processos.
    VERSÃO ESPECIAL: Tenta migrar processos sem CPF/CNPJ selecionando a opção 'Sem CPF'.
    """
    
    # Abrir diálogo para selecionar arquivo Excel ANTES de acessar o navegador
    logger.info("Selecione o arquivo Excel com a lista de processos (Migração Sem CPF)...")
    arquivo_excel = filedialog.askopenfilename(
        title="Selecione o arquivo Excel - Migração Sem CPF",
        filetypes=[("Arquivos Excel", "*.xlsx *.xls"), ("Todos os arquivos", "*.*")]
    )
    
    if not arquivo_excel:
        logger.info("Nenhum arquivo selecionado. Operação cancelada.")
        return
    
    try:
        # Ler o arquivo Excel com openpyxl
        logger.info(f"Lendo arquivo: {arquivo_excel}")
        wb = load_workbook(arquivo_excel)
        ws = wb['Planilha1']
        
        # Encontrar as colunas 'Processo' e 'Status' na primeira linha
        coluna_processo = None
        coluna_status = None

        for col_idx, cell in enumerate(ws[1], 1):
            if cell.value:
                valor_celula = str(cell.value).strip()
                if valor_celula == 'Processo':
                    coluna_processo = col_idx
                elif valor_celula == 'Status':
                    coluna_status = col_idx

        if coluna_processo is None:
            logger.info("Erro: Coluna 'Processo' não encontrada no arquivo Excel.")
            wb.close()
            return

        # Se a coluna Status não existir, criar na 2ª coluna
        if coluna_status is None:
            coluna_status = 2
            ws.cell(row=1, column=coluna_status, value="Status")
            wb.save(arquivo_excel)
            logger.info(f"✅ Coluna 'Status' criada na coluna {coluna_status}.")

        # Obter lista de processos (ignorando a primeira linha - cabeçalho)
        dados_processos = []
        for row_idx, row in enumerate(ws.iter_rows(min_row=2), start=2):
            valor_processo = row[coluna_processo - 1].value
            if valor_processo:
                proc_str = str(valor_processo).strip()
                if '/' in proc_str:
                    proc_str, incidente_str = proc_str.rsplit('/', 1)
                    incidente_str = incidente_str.strip() or '0'
                else:
                    incidente_str = '0'
                dados_processos.append({
                    'processo': proc_str,
                    'incidente': incidente_str,
                    'linha': row_idx
                })
        
        total_processos = len(dados_processos)
        logger.info(f"Total de processos encontrados: {total_processos}")
        
        # AGORA verifica se o navegador está pronto
        driver = obter_driver()
        if driver is None:
            logger.info("Navegador não está pronto. Acesse o e-Proc primeiro.")
            wb.close()
            return
        
        # Minimizar a janela do navegador após início da execução
        minimizar_navegador()
        
        # Clicar no atalhoRapido_2
        try:
            selector = '#atalhoRapido_2'
            logger.info("Clicando em 'atalhoRapido_2'...")
            elemento = WebDriverWait(driver, 5).until(
                EC.element_to_be_clickable((By.CSS_SELECTOR, selector))
            )
            elemento.click()
            logger.info("Elemento 'atalhoRapido_2' clicado com sucesso.")
        except Exception as e:
            logger.error(f"❌ ERRO: Atalho de migração não encontrado no e-Proc. Verifique se está na página correta. ({e})")
            wb.close()
            return
        
        # Iterar pelos processos
        for i, dados in enumerate(dados_processos, 1):
            if _cancelado():
                logger.info("⚠️ Execução cancelada pelo usuário.")
                break
            processo = dados['processo']
            incidente = dados['incidente']
            linha = dados['linha']
            logger.info(f"\n--- [{i}/{total_processos}] Processando: {processo} (Incidente: {incidente}) ---")
            
            try:
                # Preencher campo do número do processo
                campo_processo = '#txtNumProcesso'
                elemento = WebDriverWait(driver, 5).until(
                    EC.presence_of_element_located((By.CSS_SELECTOR, campo_processo))
                )
                elemento.clear()
                elemento.send_keys(processo)
                logger.info(f"Número do processo '{processo}' inserido.")
                
                # Preencher campo do número do incidente
                campo_incidente = '#txtNumSeqProcesso'
                elemento = driver.find_element(By.CSS_SELECTOR, campo_incidente)
                elemento.clear()
                elemento.send_keys(incidente)
                logger.info(f"Número do incidente '{incidente}' inserido.")
                
                # Clicar no botão "Buscar Processo para Migração"
                botao_buscar = '#btnBuscar'
                elemento = driver.find_element(By.CSS_SELECTOR, botao_buscar)
                elemento.click()
                logger.info(f"Botão 'Buscar Processo para Migração' clicado.")
                for _ in range(10):
                    if _cancelado():
                        break
                    time.sleep(0.1)
                if _cancelado():
                    logger.info("Cancelado após clique em Buscar.")
                    break
                
                # Aguardar a seção "Dados do Processo" carregar
                logger.info("Aguardando carregamento dos dados do processo (até 5 minutos)...")
                try:
                    # Aguardar o elemento "Dados do Processo" aparecer
                    WebDriverWait(driver, 300).until(
                        EC.visibility_of_element_located((By.XPATH, '//legend[contains(text(), "Dados do Processo")]'))
                    )
                    logger.info("✅ Dados do processo carregados.")
                    for _ in range(10):
                        if _cancelado():
                            break
                        time.sleep(0.1)
                    if _cancelado():
                        logger.info("Cancelado após carregar dados do processo.")
                        break
                    
                    # Agora verificar se tem botão Migrar ou mensagem de erro
                    tem_botao_migrar = False
                    tem_erro = False
                    tem_pessoas_sem_cpf = False
                    
                    try:
                        elemento = driver.find_element(By.CSS_SELECTOR, '#btnMigrar')
                        tem_botao_migrar = elemento.is_displayed()
                    except:
                        pass
                    
                    # Detectar erros pela extração direta dos códigos na seção de erros
                    erros_lista = extrair_codigo_erro(driver)
                    tem_erro = len(erros_lista) > 0
                    
                    try:
                        elemento = driver.find_element(By.XPATH, '//*[contains(text(), "Pessoas sem CPF/CNPJ")]')
                        tem_pessoas_sem_cpf = elemento.is_displayed()
                    except:
                        pass
                    
                    # Decidir o que fazer baseado no que encontrou
                    if tem_pessoas_sem_cpf:
                        # Encontrado "Pessoas sem CPF/CNPJ" - TENTAR TRATAMENTO AUTOMÁTICO
                        logger.info("⚠️ Processo com partes sem CPF. Tentando tratamento automático...")
                        
                        # Tenta selecionar "Sem CPF" nos dropdowns
                        selecionou = tratar_pessoas_sem_cpf(driver)
                        
                        if selecionou:
                            logger.info("✅ Opções 'Sem CPF' selecionadas. Verificando botão Migrar...")
                            # Recarregar estado dos botões após seleção
                            time.sleep(2)
                            
                            # Verificar novamente se o botão Migrar ficou disponível
                            try:
                                elemento = driver.find_element(By.CSS_SELECTOR, '#btnMigrar')
                                tem_botao_migrar = elemento.is_displayed()
                            except:
                                tem_botao_migrar = False
                            
                            # Verificar novamente se ainda há erros impeditivos
                            erros_lista = extrair_codigo_erro(driver)
                            tem_erro = len(erros_lista) > 0
                            
                            if tem_botao_migrar and not tem_erro:
                                logger.info("✅ Tratamento bem-sucedido! Prosseguindo com a migração...")
                                # Vai para o bloco de migração abaixo
                            else:
                                logger.info("❌ Ainda há erros impeditivos mesmo após selecionar 'Sem CPF'")
                                if erros_lista:
                                    descricoes = []
                                    for codigo, mensagem in erros_lista:
                                        descricao = mensagem if mensagem else obter_descricao_erro(codigo)
                                        descricoes.append(f"Erro {codigo}: {descricao}")
                                        logger.info(f"   Código identificado: {codigo} - {descricao}")
                                    status_texto = "Sem CPF Tratado - " + "; ".join(descricoes)
                                else:
                                    status_texto = "Sem CPF Tratado - Ainda com Erros"
                                
                                ws.cell(row=linha, column=coluna_status, value=status_texto)
                                wb.save(arquivo_excel)
                                logger.info(f"Status atualizado na planilha: {status_texto}")
                                
                                # Próximo processo
                                logger.info("Clicando novamente em 'atalhoRapido_2' para o próximo processo...")
                                elemento = WebDriverWait(driver, 5).until(
                                    EC.element_to_be_clickable((By.CSS_SELECTOR, '#atalhoRapido_2'))
                                )
                                elemento.click()
                                continue
                        else:
                            logger.info("❌ Não foi possível selecionar as opções automaticamente.")
                            ws.cell(row=linha, column=coluna_status, value="Não Migrado - Falha ao Tratar SEM CPF")
                            wb.save(arquivo_excel)
                            logger.info(f"Status atualizado na planilha: Não Migrado - Falha ao Tratar SEM CPF")
                            
                            # Próximo processo
                            logger.info("Clicando novamente em 'atalhoRapido_2' para o próximo processo...")
                            elemento = WebDriverWait(driver, 5).until(
                                EC.element_to_be_clickable((By.CSS_SELECTOR, '#atalhoRapido_2'))
                            )
                            elemento.click()
                            continue
                    
                    # Se chegou aqui com botão Migrar disponível, prossegue com a migração
                    if tem_botao_migrar:
                        logger.info("✅ Botão 'Migrar Processo' encontrado.")
                        
                        # Executar a função JavaScript diretamente
                        logger.info("Executando função 'importarNovo()'...")
                        driver.execute_script("importarNovo()")
                        
                        # Aguardar 2 segundos antes de aceitar o alert
                        logger.info("Aguardando caixa de diálogo de confirmação...")
                        for _ in range(20):
                            if _cancelado():
                                break
                            time.sleep(0.1)
                        if _cancelado():
                            logger.info("Cancelado antes de aceitar o alerta.")
                            break
                        
                        # Aceitar o alert JavaScript
                        alert = driver.switch_to.alert
                        alert.accept()
                        logger.info("✅ Alert JavaScript aceito.")
                        for _ in range(10):
                            if _cancelado():
                                break
                            time.sleep(0.1)
                        if _cancelado():
                            logger.info("Cancelado após aceitar o alerta.")
                            break
                        
                        # Aguardar resultado (sucesso ou erros)
                        logger.info("Aguardando resultado da migração (sucesso ou erros)...")
                        # timeout: 180 (3 min), 300 (5 min), 540 (9 min) - configurável em TIMEOUT_MIGRACAO
                        status_resultado, erros_pos = aguardar_sucesso_ou_erros(driver, timeout=TIMEOUT_MIGRACAO)

                        if status_resultado == "sucesso":
                            logger.info("✅ Processo migrado com sucesso!")
                            ws.cell(row=linha, column=coluna_status, value="Migrado com Sucesso")
                            wb.save(arquivo_excel)
                            logger.info(f"Status atualizado na planilha: Migrado com Sucesso")
                            
                            logger.info("Clicando novamente em 'atalhoRapido_2' para o próximo processo...")
                            elemento = WebDriverWait(driver, 5).until(
                                EC.element_to_be_clickable((By.CSS_SELECTOR, '#atalhoRapido_2'))
                            )
                            elemento.click()
                        elif status_resultado == "erro":
                            logger.info("❌ Erros pós-migração detectados.")
                            descricoes = []
                            for codigo, mensagem in erros_pos:
                                descricao = mensagem if mensagem else obter_descricao_erro(codigo)
                                descricoes.append(f"Erro {codigo}: {descricao}")
                                logger.info(f"   Código identificado: {codigo} - {descricao}")
                            status_texto = "; ".join(descricoes) if descricoes else "Não Migrado - Com Erros"
                            
                            ws.cell(row=linha, column=coluna_status, value=status_texto)
                            wb.save(arquivo_excel)
                            logger.info(f"Status atualizado na planilha: {status_texto}")
                            
                            # Tentar re-clicar no botão Migrar para forçar refresh
                            try:
                                logger.info("Reclicando no botão 'Migrar Processo' para forçar refresh...")
                                botao_migrar_refresh = WebDriverWait(driver, 5).until(
                                    EC.element_to_be_clickable((By.CSS_SELECTOR, '#btnMigrar'))
                                )
                                botao_migrar_refresh.click()
                                try:
                                    WebDriverWait(driver, 5).until(EC.alert_is_present())
                                    driver.switch_to.alert.accept()
                                    logger.info("Alerta de confirmação pós-erro aceito.")
                                except TimeoutException:
                                    pass
                                for _ in range(10):
                                    if _cancelado():
                                        break
                                    time.sleep(0.1)
                                try:
                                    elemento = WebDriverWait(driver, 5).until(
                                        EC.element_to_be_clickable((By.CSS_SELECTOR, '#atalhoRapido_2'))
                                    )
                                    elemento.click()
                                except Exception:
                                    pass
                            except Exception:
                                logger.info("Não foi possível re-clicar em 'Migrar Processo'. Voltando pelo atalho...")
                                try:
                                    elemento = WebDriverWait(driver, 5).until(
                                        EC.element_to_be_clickable((By.CSS_SELECTOR, '#atalhoRapido_2'))
                                    )
                                    elemento.click()
                                except Exception:
                                    pass
                        elif status_resultado == "excecao":
                            logger.info(f"❌ Exceção do sistema detectada: {erros_pos}")
                            ws.cell(row=linha, column=coluna_status, value=f"ERRO SISTEMA: {erros_pos}")
                            wb.save(arquivo_excel)
                            logger.info(f"Status atualizado na planilha: ERRO SISTEMA: {erros_pos}")
                            try:
                                elemento = WebDriverWait(driver, 5).until(
                                    EC.element_to_be_clickable((By.CSS_SELECTOR, '#atalhoRapido_2'))
                                )
                                elemento.click()
                            except Exception:
                                pass
                            continue
                        elif status_resultado == "timeout":
                            logger.info("❌ TIMEOUT: Sistema não respondeu em 5 minutos após migração.")
                            logger.info("⚠️ Tentando recuperar e continuar com o próximo processo...")
                            ws.cell(row=linha, column=coluna_status, value="ERRO SISTEMA: Timeout após migração")
                            wb.save(arquivo_excel)
                            logger.info("⚠️ Possível problema de conexão, sistema lento ou travado.")

                            if _tentar_recuperar_navegador(driver):
                                continue
                            else:
                                logger.info("❌ Falha na recuperação.")
                                logger.info("⚠️ EXECUÇÃO INTERROMPIDA - Não foi possível recuperar o navegador.")
                                break
                        else:  # cancelado
                            logger.info("Cancelado durante a espera pelo resultado da migração.")
                            break

                    elif tem_erro:
                        # Erros de validação encontrados via extração direta
                        logger.info("❌ Erros de validação encontrados.")
                        if erros_lista:
                            descricoes = []
                            for codigo, mensagem in erros_lista:
                                descricao = mensagem if mensagem else obter_descricao_erro(codigo)
                                descricoes.append(f"Erro {codigo}: {descricao}")
                                logger.info(f"   Código identificado: {codigo} - {descricao}")
                            status_texto = "; ".join(descricoes)
                        else:
                            status_texto = "Não Migrado - Com Erros"

                        ws.cell(row=linha, column=coluna_status, value=status_texto)
                        wb.save(arquivo_excel)
                        logger.info(f"Status atualizado na planilha: {status_texto}")

                        logger.info("Clicando novamente em 'atalhoRapido_2' para o próximo processo...")
                        elemento = WebDriverWait(driver, 5).until(
                            EC.element_to_be_clickable((By.CSS_SELECTOR, '#atalhoRapido_2'))
                        )
                        elemento.click()
                        continue

                    else:
                        # Nem botão nem erro encontrado
                        logger.info("⚠️ Nem botão Migrar nem mensagem de erro foram encontrados.")

                        ws.cell(row=linha, column=coluna_status, value="Não Migrado - Com Erros")
                        wb.save(arquivo_excel)
                        logger.info(f"Status atualizado na planilha: Não Migrado - Com Erros")
                        
                        logger.info("Clicando novamente em 'atalhoRapido_2' para o próximo processo...")
                        try:
                            elemento = WebDriverWait(driver, 5).until(
                                EC.element_to_be_clickable((By.CSS_SELECTOR, '#atalhoRapido_2'))
                            )
                            elemento.click()
                        except:
                            pass
                        continue
                    
                except TimeoutException as e:
                    logger.info(f"❌ TIMEOUT: Dados do processo não carregaram em 5 minutos.")
                    logger.info(f"⚠️ Tentando recuperar e continuar com o próximo processo...")
                    logger.info(f"⚠️ Detalhes: {e}")
                    
                    ws.cell(row=linha, column=coluna_status, value="ERRO SISTEMA: Timeout ao carregar dados")
                    wb.save(arquivo_excel)
                    
                    if _tentar_recuperar_navegador(driver):
                        continue
                    else:
                        logger.info("❌ Falha na recuperação.")
                        logger.info("⚠️ EXECUÇÃO INTERROMPIDA - Não foi possível recuperar o navegador.")
                        break
                    
                except Exception as e:
                    logger.info(f"❌ ERRO ao processar dados do processo: {e}")
                    logger.info(f"⚠️ Tentando recuperar e continuar com o próximo processo...")
                    
                    ws.cell(row=linha, column=coluna_status, value="ERRO SISTEMA: Erro ao processar dados")
                    wb.save(arquivo_excel)
                    
                    if _tentar_recuperar_navegador(driver):
                        continue
                    else:
                        logger.info("❌ Falha na recuperação.")
                        logger.info("⚠️ EXECUÇÃO INTERROMPIDA - Não foi possível recuperar o navegador.")
                        break
                
            except Exception as e:
                logger.info(f"❌ ERRO ao processar o processo {processo}: {e}")
                logger.info(f"⚠️ Tentando recuperar e continuar com o próximo processo...")
                logger.info(f"⚠️ Possíveis causas: elementos não encontrados, conexão temporária perdida.")
                
                ws.cell(row=linha, column=coluna_status, value="ERRO SISTEMA: Falha ao processar")
                wb.save(arquivo_excel)
                
                if _tentar_recuperar_navegador(driver):
                    continue
                else:
                    logger.info("❌ Falha na recuperação.")
                    logger.info("⚠️ EXECUÇÃO INTERROMPIDA - Não foi possível recuperar o navegador.")
                    break
        
        # Fechar a planilha no final
        wb.close()
        logger.info(f"\n✅ Processamento concluído! Planilha atualizada: {arquivo_excel}")
            
    except Exception as e:
        logger.info(f"Erro ao processar arquivo Excel: {e}")